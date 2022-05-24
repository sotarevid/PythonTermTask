from openpyxl import load_workbook
import os
from copy import copy
import datetime
from collections import Counter


class Staff:
    """
    Staf constractor
    """

    def __init__(self, name, position, bid, days_off):
        """

        :param name: Name of user
        :param position: Position of user
        :param bid: Bid of user
        :param days_off: User's daysoff
        """
        self.name = name
        self.position = position
        self.bid = bid
        self.days_off = days_off


def add_weekends_data(row, weekends, ws):
    """
    Fills user weekends

    :param ws: worksheet
    :param row: row for filling
    :param weekends: list of weekends
    """
    row_weekends = ws["E16":"AQ16"]

    for cell in range(row_weekends[0].__len__()):
        if row_weekends[0][cell].value in weekends:
            row[0][cell].value = "В"


def move_template_data(n, ws):
    """
    Moves static data from template. IDK how it works, but it works

    :param ws: worksheet
    :param n: count of added users(lines)
    """

    ws.merge_cells(start_row=21 + n, start_column=13,
                   end_row=21 + n, end_column=16)
    ws.unmerge_cells("M20:P20")
    ws.merge_cells(start_row=21 + n, start_column=21,
                   end_row=21 + n, end_column=25)
    ws.unmerge_cells("U20:Y20")
    ws.merge_cells(start_row=21 + n, start_column=27,
                   end_row=21 + n, end_column=31)
    ws.unmerge_cells("AA20:AE20")
    ws.merge_cells(start_row=21 + n, start_column=35,
                   end_row=22 + n, end_column=37)
    ws.unmerge_cells("AI20:AK21")
    ws.merge_cells(start_row=20 + n, start_column=38,
                   end_row=20 + n, end_column=43)
    ws.unmerge_cells("AL19:AQ19")


def format_cells(n, ws):
    """
    Format cells in row

    :param ws: worksheet
    :param n: row number
    """
    example_row = ws.row_dimensions[n]
    ws.row_dimensions[n + 1] = example_row


def ins_rows(count, ws):
    """
    Insert rows, format and styles them

    :param ws: worksheet
    :param count: number of rows to insert
    """
    move_template_data(count - 2, ws)
    for s in range(count - 1):
        ws.insert_rows(17, amount=1)
        cells = ws['A18': 'AQ18']
        new_cells = ws['A17': 'AQ17']
        cell_style(cells, new_cells)
        format_cells(17 + s, ws)
        format_cells(20 + s, ws)


def add_dayoffs(row, day_offs, ws):
    """
    Add staf's dayoffs to row

    :param row: row of user workdays
    :param day_offs: staf dayoffs
    :param ws: worksheet
    """
    row_month = ws["E16":"AQ16"]

    for cell in range(row_month[0].__len__()):
        for day in day_offs:
            if row_month[0][cell].value in day:
                row[0][cell].value = day[1]


def count_workdays(row):
    """
    Count workdays

    :param row: row of user workdays
    :return: count of workdays
    """
    # Формула из ячейки АК17 =ЦЕЛОЕ(РАЗНДАТ($AF$9;$AI$9;"d")+ 1)-СЧЁТЕСЛИ(E17:AI17;"В")-AM17-AO17-AQ17
    count = 0
    for cell in range(row[0].__len__()):
        if row[0][cell].value is None:
            count += 1
    return count


def find_causes(row):
    """
    Find reasons for absence.

    :param row:row of user workdays
    :return: list of causes
    """
    causes = []
    for cell in range(row[0].__len__()):
        if (row[0][cell].value != "ОТ") & (row[0][cell].value != "В") & (row[0][cell].value != "Б") & (
                row[0][cell].value is not None):
            causes.append(row[0][cell].value)
    return causes


def fill_absence_reasons(ws, row, count):
    """
    Fills absence reasons

    :param ws: worksheet
    :param row:row of user workdays
    :param count: num of current line
    :return:
    """
    causes = find_causes(row)
    causes_counter = Counter(causes)
    if causes_counter.__len__() == 0:
        pass
    else:
        causes_count = list(dict(causes_counter).values())
        causes_code = list(dict(causes_counter).keys())
        ws["AN" + str(16 + count)].value = causes_code[0]
        ws["AP" + str(16 + count)].value = causes_code[1]
        ws["AO" + str(16 + count)].value = causes_count[0]
        ws["AQ" + str(16 + count)].value = causes_count[1]


def write_data_to_file(data, weekends, ws):
    """
    Write data to worksheet

    :param ws: worksheet
    :param weekends: list of weekends
    :param data: DB data
    """
    ws['AF9'].value = data[2]
    count = 0
    ins_rows(data[0].__len__(), ws)
    for s in data[0]:
        count += 1
        ws["A" + str(16 + count)].value = count
        ws['B' + str(16 + count)].value = s.name
        ws['C' + str(16 + count)].value = s.position
        ws['D' + str(16 + count)].value = s.bid
        row = ws["E" + str(16 + count) + ":AI" + str(16 + count)]
        add_dayoffs(row, s.days_off, ws)
        add_weekends_data(row, weekends, ws)
        ws["AJ" + str(16 + count)].value = "01"
        ws["AK" + str(16 + count)].value = count_workdays(row)
        fill_absence_reasons(ws, row, count)


def cell_style(cells, new_cells):
    """
    Change the style of cell to example cell style

    :param cells: range of example cells style
    :param new_cells: range of cells to be styled
    :return: range styled cells as new_cells
    """
    for c in range(cells[0].__len__()):
        if cells[0][c].has_style:
            new_cells[0][c].font = copy(cells[0][c].font)
            new_cells[0][c].border = copy(cells[0][c].border)
            new_cells[0][c].fill = copy(cells[0][c].fill)
            new_cells[0][c].number_format = copy(cells[0][c].number_format)
            new_cells[0][c].protection = copy(cells[0][c].protection)
            new_cells[0][c].alignment = copy(cells[0][c].alignment)
    return new_cells


def save_file(date, wb_template):
    """
    Save file as NAME

    :param wb_template: template
    :param date: the reporting date
    :return: new file
    """
    table_filename = 'Производственный календарь {}.xlsx'.format(date)
    wb_template.save(table_filename)
    return table_filename


def generate(date=None, weekends=None, staf=None):
    """
    Main function to create xlsx file

    :param date: month of unloading in format datetime.date(YYYY, MM, DD)
    :param weekends: list of weekends
    :param staf: list of objects from class Staf
    """
    os.chdir(r"xlsx-table")  # переходим в текущую дерикторию иначе блядь не работает
    wb_template = load_workbook(filename='template.xlsx')  # rename to Sample
    ws = wb_template.active
    if date is None:
        date = datetime.date(2022, 1, 1)
    if staf is None:
        staf = [
            Staff('Dick', 'Big', 0.5, [
                  (2, "ОТ"), (3, 'ОТ'), (4, "ОТ"), (25, "Б"), (14, "Гы"), (15, 'Гы'), (17, 'ЗЩ')]),
            Staff('Lol', 'kek', 1, [
                (2, "Б"), (5, 'ОТ'), (6, "ОТ"), (27, "Б")]),
            Staff('Yaga', 'Baba', 0.33,
                  [(3, "ОТ"), (5, 'ОТ'), (9, "Б"), (22, "РР"), (23, "ТТ"), (26, "ТТ"), (27, 'РР')]),
            Staff('Dad', 'Papa', 2,
                  [(12, "Б"), (13, 'Б'), (22, "РР"), (23, "II"), (24, "NN"), (28, "ОТ"), (29, "ОТ"), (30, 'ОТ')])]
    if weekends is None:
        weekends = [1, 2, 12, 13, 20, 21, 28, 31]
    data = [staf, weekends, date]
    write_data_to_file(data, weekends, ws)
    return save_file(date, wb_template)


if __name__ == '__main__':
    generate()
