import os

from openpyxl import Workbook
from openpyxl import load_workbook

os.remove('табель Август 20.xlsx')
os.remove('data_.xlsx')
year = 2022
mounth_num = 8
from lxml import html
import requests

page = requests.get("http://www.consultant.ru/law/ref/calendar/proizvodstvennye/{}/".format(year))

tree = html.fromstring(page.content)
month_tags = tree.xpath("//th[@class='month']/../../..")
holidays_by_month = []
holidays_count = 0
for t in month_tags:
    holidays = t.xpath(
        ".//td[@class='holiday weekend' or @class='weekend' or @class='nowork']/text()"
    )
    holidays_list = [int(day) for day in holidays]
    holidays_count += len(holidays_list)
    holidays_by_month.append(holidays_list)
print(holidays_by_month)
mounths = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь',
           'Декабрь']
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! вывод выходных
wb = Workbook()
dest_filename = 'data_.xlsx'
ws1 = wb.active
ws1.title = "выходные"
_ = ws1.cell(column=1, row=1, value="номер месяца")
_ = ws1.cell(column=2, row=1, value="месяц")
_ = ws1.cell(column=3, row=1, value="номер выходного дня")
row = 1
for m_num, holidays in enumerate(holidays_by_month):
    for h in holidays:
        row += 1
        _ = ws1.cell(column=1, row=row, value="{}".format(m_num + 1))
        _ = ws1.cell(column=2, row=row, value="{}".format(mounths[m_num]))
        _ = ws1.cell(column=3, row=row, value="{}".format(h))

wb.save(filename=dest_filename)
import calendar
import datetime

days_in_mounth = calendar.monthrange(year, mounth_num)[1]
print(days_in_mounth)
wb_data = load_workbook(filename='../../../xlsx-data/data.xlsx')
sheet_ranges = wb_data['отпуска больничные и пр']
print(type(sheet_ranges.cell(column=2, row=2).value))

holidays_ws = wb_data['выходные']
holidays_list = []
for i in range(2, holidays_ws.max_row + 1):
    if int(holidays_ws.cell(column=1, row=i).value) == mounth_num:
        holidays_list.append(int(holidays_ws.cell(column=3, row=i).value))

print(holidays_list)

working_day_10 = 10
while working_day_10 in holidays_list:
    working_day_10 += 1

working_day_20 = 20
while working_day_20 in holidays_list:
    working_day_20 += 1

print(working_day_10, working_day_20)

# class 'datetime.datetime'


# [1, 2, 8, 9, 15, 16, 22, 23, 29, 30]
# 10
# 20


staff_ws = wb_data['сотрудники']
number_of_staff = staff_ws.max_row - 1
absence_dict = {}
for n in range(number_of_staff):
    absence_dict[staff_ws.cell(column=2, row=n + 2).value] = []
print(number_of_staff)
# {'Вылков А.И.': [], 'Зверев В.С.': []})
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!11
#wb = Workbook()
dest_filename = 'табель {} {}.xlsx'.format(mounths[mounth_num - 1], working_day_20)

wb_template = load_workbook(filename='template.xlsx')
sheet_name = str(days_in_mounth)

for name in wb_template.sheetnames:
    ws = wb_template[name]
    if name != sheet_name:
        wb_template.remove(ws)

ws = wb_template[sheet_name]
ws.title = 'табель'

for k in range(number_of_staff - 1):
    ws.insert_rows(17)

staff_ws = wb_data['сотрудники']
for k in range(number_of_staff):
    for c in range(1, 5):
        ws.cell(row=17 + k, column=c).value = staff_ws.cell(row=2 + k, column=c).value
    for h in holidays_list:
        ws.cell(row=17 + k, column=h + 4).value = 'В'

another_label_ws = wb_data['отпуска больничные и пр']
r = 2
value = another_label_ws.cell(row=r, column=1).value

while value:
    start_date = another_label_ws.cell(row=r, column=2).value.date()
    end_date = another_label_ws.cell(row=r, column=3).value.date()
    letter_label = another_label_ws.cell(row=r, column=4).value
    digit_label = another_label_ws.cell(row=r, column=5).value

    for k in range(number_of_staff):
        days_absense = 0
        if value == staff_ws.cell(row=2 + k, column=2).value:
            for d in range(1, days_in_mounth + 1):
                if start_date <= datetime.date(year, mounth_num, d) <= end_date:
                    ws.cell(row=17 + k, column=d + 4).value = letter_label
                    days_absense += 1
            if days_absense > 0:
                ws.cell(row=17 + k, column=4 + days_in_mounth + 5).value = digit_label
                ws.cell(row=17 + k, column=4 + days_in_mounth + 6).value = days_absense
                absence_dict[value].append([days_absense, letter_label])
    r += 1
    value = another_label_ws.cell(row=r, column=1).value

ws.move_range("O{}:AQ{}".format(ws.max_row - 2, ws.max_row), rows=20, cols=0)
ws.delete_rows(17 + number_of_staff, 1)

ws.merge_cells(start_row=17 + number_of_staff + 2, start_column=15, end_row=17 + number_of_staff + 2, end_column=25)
ws['Y9'] = datetime.date(year, mounth_num, working_day_20)

wb_template.save(dest_filename)
